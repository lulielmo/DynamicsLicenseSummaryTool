#!/usr/bin/env python3

import pandas as pd
import re
import os
import sys
import argparse
from collections import Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path

def log(message, verbose=False, always=False):
    """
    Utility function to handle logging messages.
    Args:
        message: The message to print
        verbose: Whether this is a debug message that should only be shown in verbose mode
        always: Whether to always show this message regardless of verbose setting
    """
    global VERBOSE
    if always or (verbose and VERBOSE):
        print(message)

# Global verbose flag
VERBOSE = False

def load_roles_from_file(roles_file):
    """
    Read roles and their license requirements from an Excel file.
    Expected format:
    Column A: Role names
    Columns B-F: License requirements (1 if required, empty if not)
    B: Finance, C: SCM, D: Commerce, E: Project, F: HR
    """
    log(f"Reading roles from file: {roles_file}", verbose=True)
    try:
        df = pd.read_excel(roles_file)
        roles = {}
        for _, row in df.iterrows():
            role = row.iloc[0]  # Role name from column A
            if pd.isna(role):
                continue
            
            # Get license requirements from columns B-F
            licenses = {
                'Finance': bool(row.iloc[1] == 1) if len(row) > 1 and not pd.isna(row.iloc[1]) else False,
                'SCM': bool(row.iloc[2] == 1) if len(row) > 2 and not pd.isna(row.iloc[2]) else False,
                'Commerce': bool(row.iloc[3] == 1) if len(row) > 3 and not pd.isna(row.iloc[3]) else False,
                'Project': bool(row.iloc[4] == 1) if len(row) > 4 and not pd.isna(row.iloc[4]) else False,
                'HR': bool(row.iloc[5] == 1) if len(row) > 5 and not pd.isna(row.iloc[5]) else False
            }
            roles[str(role)] = licenses
        
        return roles
    except Exception as e:
        log(f"Error reading roles file: {e}", always=True)
        return {}

def extract_roles(excel_file, roles_file):
    """
    Extract role combinations and their counts from the Excel file.
    Requires a roles file to specify which roles to look for.
    """
    try:
        # Read roles from file
        target_roles = load_roles_from_file(roles_file)
        if not target_roles:
            log("Error: No roles found in the roles file. Please check the file format.", always=True)
            return [], {}, {}
            
        log(f"Found {len(target_roles)} roles in roles file", always=True)
        
        log("\nAvailable roles:", verbose=True)
        for role in target_roles:
            log(f"- {role}", verbose=True)
        
        log("\nAnalyzing file...", always=True)
        # Read the Excel file using openpyxl to handle locked rows
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        sheet = wb.active
        
        # Convert to pandas DataFrame for easier processing, skipping the first 19 rows
        data = []
        row_count = 0
        for row in sheet.iter_rows(min_row=20, values_only=True):
            row_count += 1
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if len(df) == 0:
            log("Error: No data found after skipping header rows", always=True)
            return [], {}, {}
        
        # Dictionary to store role combinations and their counts
        role_counts = {}
        # Dictionary to store license requirements for each combination
        license_requirements = {}
        # Dictionary to store role combination types
        combination_types = {}
        
        # Dictionary to track unique users and their roles
        user_roles = {}
        current_user = None
        
        # Process each row
        i = 0
        while i < len(df):
            row = df.iloc[i]
            
            # Check if this is a user header row (Alias is in column 3)
            if isinstance(row.iloc[3], str) and row.iloc[3].strip() == "Alias":
                log(f"\nFound Alias header at row {i}", verbose=True)
                # User data is in the next row
                if i + 1 < len(df):
                    user_row = df.iloc[i + 1]
                    current_user = str(user_row.iloc[3])
                    if current_user not in user_roles:
                        user_roles[current_user] = set()
                        log(f"Found user: {current_user}", verbose=True)
                    
                    # Skip to security role headers (2 rows down)
                    i += 2
                    if i < len(df):
                        role_header_row = df.iloc[i]
                        if isinstance(role_header_row.iloc[5], str) and role_header_row.iloc[5].strip() == "Security Role":
                            log("Found Security Role header", verbose=True)
                            # Read roles from next row(s) until we hit another user or end of file
                            i += 1
                            while i < len(df):
                                role_row = df.iloc[i]
                                # Check if we've hit another user section
                                if isinstance(role_row.iloc[3], str) and role_row.iloc[3].strip() == "Alias":
                                    i -= 1  # Back up one row so we process this header in the next iteration
                                    break
                                
                                # Process role if it exists
                                if isinstance(role_row.iloc[5], str):
                                    security_role = str(role_row.iloc[5])
                                    if not pd.isna(security_role) and security_role != "nan":
                                        # Split roles and filter based on target roles
                                        roles = [r.strip() for r in security_role.split(',')]
                                        log(f"Found roles for {current_user}: {roles}", verbose=True)
                                        matching_roles = [r for r in roles if r in target_roles]
                                        if matching_roles:
                                            log(f"Matching roles: {matching_roles}", verbose=True)
                                            user_roles[current_user].update(matching_roles)
                                i += 1
            i += 1
        
        log("\nFound users with roles:", verbose=True)
        for user, roles in user_roles.items():
            if roles:
                log(f"{user}: {roles}", verbose=True)
        
        # Process unique role combinations for each user
        for user, roles in user_roles.items():
            if roles:
                # Sort roles for consistent combination strings
                role_list = sorted(roles)
                role_combination = ' + '.join(role_list)
                
                # Count occurrences
                role_counts[role_combination] = role_counts.get(role_combination, 0) + 1
                
                # Calculate license requirements for the combination
                combined_licenses = {
                    'Finance': False,
                    'SCM': False,
                    'Commerce': False,
                    'Project': False,
                    'HR': False
                }
                
                # Combine license requirements
                for role in role_list:
                    role_licenses = target_roles[role]
                    for license_type in combined_licenses:
                        combined_licenses[license_type] |= role_licenses[license_type]
                
                license_requirements[role_combination] = combined_licenses
                
                # Create combination types based on required licenses
                combination_type = []
                if combined_licenses['Finance']: combination_type.append('Finance')
                if combined_licenses['SCM']: combination_type.append('SCM')
                if combined_licenses['Commerce']: combination_type.append('Commerce')
                if combined_licenses['Project']: combination_type.append('Project')
                if combined_licenses['HR']: combination_type.append('HR')
                
                type_str = ', '.join(combination_type)
                if type_str not in combination_types:
                    combination_types[type_str] = {}
                combination_types[type_str][role_combination] = role_counts[role_combination]
        
        # Sort results by count (descending)
        sorted_combinations = sorted(role_counts.items(), key=lambda x: x[1], reverse=True)
        
        log(f"\nFound {len(role_counts)} unique role combinations", always=True)
        log(f"Total users with matching roles: {sum(role_counts.values())}", always=True)
        
        wb.close()  # Close the workbook
        return sorted_combinations, license_requirements, combination_types
        
    except Exception as e:
        log(f"Error processing Excel file: {e}", always=True)
        log(f"Error details: {e.__class__.__name__}", always=True)
        import traceback
        traceback.print_exc()
        return [], {}, {}

def create_output_filename(input_file):
    """
    Create output filename by adding '_summary' before the extension
    """
    input_path = Path(input_file)
    return str(input_path.parent / f"{input_path.stem}_summary{input_path.suffix}")

def write_results_to_excel_file(results, output_file):
    """
    Write results to a new Excel file.
    """
    wb = Workbook()
    sheet = wb.active
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # Check if we have license information
    has_license_info = bool(results[1]) if len(results) > 1 else False
    has_license_combinations = bool(results[2]) if len(results) > 2 else False
    
    # Prepare headers
    headers = [
        ("A1", "Count", 10),
        ("B1", "Role Combination", 40)
    ]
    
    # Add license requirement headers if available
    if has_license_info:
        headers.extend([
            ("C1", "Finance", 10),
            ("D1", "SCM", 10),
            ("E1", "Commerce", 10),
            ("F1", "Project", 10),
            ("G1", "HR", 10)
        ])
    
    # Add empty column before license combinations and combination headers
    if has_license_combinations:
        headers.append(("H1", "", 15))  # Empty column (H)
        
        # Get ordered combinations for consistent column ordering
        ordered_combinations = sorted(results[2].keys())
        
        # Add license combination headers with correct alignment
        for i, combo_type in enumerate(ordered_combinations):
            col_letter = chr(ord('I') + i)  # Start from I onwards
            headers.append((f"{col_letter}1", combo_type, 20))
    
    # Write and format headers
    for cell_ref, value, width in headers:
        cell = sheet[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        sheet.column_dimensions[cell_ref[0]].width = width
    
    # Write data
    sorted_combinations, license_requirements, combination_types = results
    for i, (combination, count) in enumerate(sorted_combinations, start=2):
        # Write count and combination
        sheet.cell(row=i, column=1, value=count).alignment = Alignment(horizontal="center")
        sheet.cell(row=i, column=2, value=combination)
        
        # Write license requirements if available
        if has_license_info and combination in license_requirements:
            licenses = license_requirements[combination]
            for col, license_type in enumerate(['Finance', 'SCM', 'Commerce', 'Project', 'HR'], start=3):
                if licenses[license_type]:
                    cell = sheet.cell(row=i, column=col, value=count)
                    cell.alignment = Alignment(horizontal="center")
        
        # Write license combinations
        if has_license_combinations:
            license_combinations = {}
            for combo_type, combinations in results[2].items():
                if combination in combinations:
                    license_combinations[combo_type] = combinations[combination]
            
            # Use same order as for headers
            for j, combo_type in enumerate(ordered_combinations):
                if combo_type in license_combinations:
                    col = ord('I') - ord('A') + j  # Convert to column number (I=8, J=9, etc.)
                    cell = sheet.cell(row=i, column=col + 1)  # +1 because Excel uses 1-based indexing
                    cell.value = license_combinations[combo_type]
                    cell.alignment = Alignment(horizontal="center")
    
    # Format summary row
    last_row = len(results[0]) + 2
    sum_font = Font(bold=True)
    sum_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    sum_alignment = Alignment(horizontal="center")
    sum_border = Border(top=Side(style="medium"))
    
    # Add "Total" in column B
    sum_cell = sheet.cell(row=last_row, column=2, value="Total")
    
    # Calculate and format sums for all relevant columns
    # Column A (Count)
    col_a_sum = sum(sheet.cell(row=i, column=1).value or 0 for i in range(2, last_row))
    sheet.cell(row=last_row, column=1, value=col_a_sum)
    
    # Columns C-G (License requirements)
    if has_license_info:
        for col in range(3, 8):  # Columns C-G
            col_sum = sum(sheet.cell(row=i, column=col).value or 0 for i in range(2, last_row))
            sheet.cell(row=last_row, column=col, value=col_sum)
    
    # Dynamic columns (I onwards)
    if has_license_combinations:
        for j, _ in enumerate(ordered_combinations):
            col = ord('I') - ord('A') + j  # Convert to column number (I=8, J=9, etc.)
            col_sum = sum(sheet.cell(row=i, column=col + 1).value or 0 for i in range(2, last_row))
            sheet.cell(row=last_row, column=col + 1, value=col_sum)
    
    # Apply formatting to all columns in summary row
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=last_row, column=col)
        cell.font = sum_font
        cell.fill = sum_fill
        cell.border = sum_border
        cell.alignment = sum_alignment
    
    # Save workbook
    wb.save(output_file)
    log(f"\nResults written to: {output_file}", always=True)

def main():
    parser = argparse.ArgumentParser(description='Analyze Dynamics 365 license requirements based on user roles.')
    parser.add_argument('excel_file', help='The Excel file containing the license report')
    parser.add_argument('roles_file', help='The Excel file containing role definitions and license requirements')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose output for debugging')
    
    args = parser.parse_args()
    
    global VERBOSE
    VERBOSE = args.verbose
    
    if not os.path.exists(args.excel_file):
        log(f"Error: File not found: {args.excel_file}", always=True)
        return
    
    if not os.path.exists(args.roles_file):
        log(f"Error: File not found: {args.roles_file}", always=True)
        return
    
    log(f"Processing file: {args.excel_file}", always=True)
    log(f"Using roles from: {args.roles_file}", always=True)
    
    results = extract_roles(args.excel_file, args.roles_file)
    if results[0]:  # If we have any results
        output_file = create_output_filename(args.excel_file)
        write_results_to_excel_file(results, output_file)
    else:
        log("No matching roles found in the file.", always=True)

if __name__ == "__main__":
    main() 